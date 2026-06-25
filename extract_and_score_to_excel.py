#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Extract image metadata and merge objective metric JSONL files into Excel.

The scoring itself can be executed in different conda environments by
objective_score.py. This script only needs openpyxl and standard libraries.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".bmp"}
FILENAME_PATTERN = re.compile(
    r"^c_"
    r"(?P<model>[^_]+)_"
    r"(?P<prompt_no>[^_]+)_"
    r"(?P<language>chinese|english)_"
    r"(?P<width>\d+)-(?P<height>\d+)_"
    r"step(?P<step>\d+)_"
    r"(?P<seed_no>\d+)"
    r"\.(?P<ext>png|jpg|jpeg|webp|bmp)$",
    re.IGNORECASE,
)

METRIC_COLUMNS = [
    "OpenCLIPScore",
    "CNCLIPScore",
    "AestheticScore",
    "MUSIQScore",
    "ImageRewardScore",
    "HPSv2.1Score",
    "VQAScore",
]

OUTPUT_HEADERS = [
    ("model", "ModelShortName"),
    ("prompt_no", "PromptID"),
    ("language", "Language"),
    ("resolution", "Resolution"),
    ("step", "Steps"),
    ("seed_no", "SeedNo"),
    *[(metric, metric) for metric in METRIC_COLUMNS],
    ("image_size_bytes", "ImageSizeBytes"),
    ("elapsed_seconds", "ElapsedSeconds"),
    ("iter_per_second", "IterPerSecond"),
    ("vram_usage", "VRAMUsage"),
    ("gpu_usage", "GPUUsage"),
    ("memory_usage", "MemoryUsage"),
    ("cpu_usage", "CPUUsage"),
    ("filename", "FileName"),
]

RESOURCE_OUTPUT_HEADERS = [
    ("filename", "FileName"),
    ("metric", "Metric"),
    ("score", "Score"),
    ("elapsed_seconds", "ElapsedSeconds"),
    ("started_at_utc", "StartedAtUTC"),
    ("ended_at_utc", "EndedAtUTC"),
    ("sample_interval_seconds", "SampleIntervalSeconds"),
    ("sample_count", "SampleCount"),
    ("process_cpu_avg", "ProcessCPUAvgPercent"),
    ("process_cpu_max", "ProcessCPUMaxPercent"),
    ("process_memory_avg", "ProcessMemoryRSSAvgMB"),
    ("process_memory_max", "ProcessMemoryRSSMaxMB"),
    ("system_cpu_avg", "SystemCPUAvgPercent"),
    ("system_cpu_max", "SystemCPUMaxPercent"),
    ("system_memory_percent_avg", "SystemMemoryAvgPercent"),
    ("system_memory_percent_max", "SystemMemoryMaxPercent"),
    ("system_memory_used_avg", "SystemMemoryUsedAvgMB"),
    ("system_memory_used_max", "SystemMemoryUsedMaxMB"),
    ("gpu_index", "GPUIndex"),
    ("gpu_name", "GPUName"),
    ("gpu_utilization_avg", "GPUUtilizationAvgPercent"),
    ("gpu_utilization_max", "GPUUtilizationMaxPercent"),
    ("gpu_memory_avg", "GPUMemoryUsedAvgMB"),
    ("gpu_memory_max", "GPUMemoryUsedMaxMB"),
    ("monitor_error", "MonitorError"),
    ("metric_error", "MetricError"),
]


def _smart_cast(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    text = value.strip()
    if text == "":
        return ""
    lower = text.lower()
    if lower in {"none", "null"}:
        return None
    if lower in {"true", "false"}:
        return lower == "true"
    if re.fullmatch(r"[+-]?\d+", text):
        try:
            return int(text)
        except Exception:
            pass
    if re.fullmatch(r"[+-]?\d+(\.\d+)?", text):
        try:
            return float(text)
        except Exception:
            pass
    return text


def _parse_key_value_text(text: str) -> Dict[str, Any]:
    data: Dict[str, Any] = {}
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        m = re.match(r"^([A-Za-z0-9_\-\u4e00-\u9fa5]+)\s*:\s*(.+)$", line)
        if not m:
            m = re.match(r"^([A-Za-z0-9_\-\u4e00-\u9fa5]+)\s*=\s*(.+)$", line)
        if m:
            key = m.group(1).strip()
            value = m.group(2).strip()
            data[key] = _smart_cast(value)
    return data


def _read_sidecar_txt(txt_path: Path) -> Dict[str, Any]:
    text = txt_path.read_text(encoding="utf-8").strip()
    if not text:
        return {}
    try:
        obj = json.loads(text)
        if isinstance(obj, dict):
            return obj
    except Exception:
        pass
    return _parse_key_value_text(text)


def _pick_first_value(data: Dict[str, Any], keys: Sequence[str], default: Any = None) -> Any:
    for key in keys:
        if key in data:
            return data[key]
    lowered = {str(k).lower(): v for k, v in data.items()}
    for key in keys:
        lk = key.lower()
        if lk in lowered:
            return lowered[lk]
    return default


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except Exception:
        return None


def _get_nested(data: Dict[str, Any], path: str, default: Any = None) -> Any:
    current: Any = data
    for part in path.split("."):
        if not isinstance(current, dict) or part not in current:
            return default
        current = current[part]
    return current


RESOURCE_KEY_CANDIDATES: Dict[str, List[str]] = {
    "vram_usage": [
        "peak_vram_mb", "max_vram_mb", "peak_vram", "max_vram", "vram_peak",
        "gpu_memory_peak", "gpu_memory_max", "vram_usage", "vram", "gpu_memory",
        "gpu_memory_usage", "VRAMPeak", "MaxVRAM", "VRAMUsage", "Info",
    ],
    "gpu_usage": [
        "avg_gpu_usage", "avg_gpu_util", "avg_gpu_utilization", "gpu_usage_avg",
        "gpu_util_avg", "gpu_avg", "gpu_usage", "gpu_util", "gpu_utilization",
        "gpu_percent", "gpu", "MeanGPUUtilization", "GPUMeanUsage", "GPUUsage",
        "GPUUtilization",
    ],
    "memory_usage": [
        "avg_memory_usage", "avg_ram_usage", "memory_usage_avg", "ram_usage_avg",
        "system_memory_avg", "memory_usage", "memory", "ram_usage", "ram",
        "system_memory_usage", "MeanMemoryUsage", "MemoryMeanUsage", "MemoryUsage", "Memory",
    ],
    "cpu_usage": [
        "avg_cpu_usage", "avg_cpu_percent", "cpu_usage_avg", "cpu_percent_avg",
        "cpu_avg", "cpu_usage", "cpu", "cpu_percent", "MeanCPUUtilization",
        "CPUMeanUsage", "CPUUsage", "CPUUtilization",
    ],
}


def _first_not_none(*values: Any) -> Any:
    for value in values:
        if value is not None:
            return value
    return None


def _extract_resource_fields(txt_info: Dict[str, Any]) -> Tuple[Any, Any, Any, Any]:
    vram_usage = _first_not_none(
        _get_nested(txt_info, "monitor.gpu.memory_used_mb.max"),
        _get_nested(txt_info, "monitor.gpu.memory_used_mb.avg"),
        _pick_first_value(txt_info, RESOURCE_KEY_CANDIDATES["vram_usage"], default=None),
    )
    gpu_usage = _first_not_none(
        _get_nested(txt_info, "monitor.gpu.utilization_percent.avg"),
        _get_nested(txt_info, "monitor.gpu.utilization_percent.max"),
        _pick_first_value(txt_info, RESOURCE_KEY_CANDIDATES["gpu_usage"], default=None),
    )
    memory_usage = _first_not_none(
        _get_nested(txt_info, "monitor.comfyui_memory_rss_mb.avg"),
        _get_nested(txt_info, "monitor.system_memory_used_mb.avg"),
        _get_nested(txt_info, "monitor.comfyui_memory_percent.avg"),
        _get_nested(txt_info, "monitor.system_memory_percent.avg"),
        _pick_first_value(txt_info, RESOURCE_KEY_CANDIDATES["memory_usage"], default=None),
    )
    cpu_usage = _first_not_none(
        _get_nested(txt_info, "monitor.comfyui_cpu_percent.avg"),
        _get_nested(txt_info, "monitor.system_cpu_percent.avg"),
        _pick_first_value(txt_info, RESOURCE_KEY_CANDIDATES["cpu_usage"], default=None),
    )
    return vram_usage, gpu_usage, memory_usage, cpu_usage


def extract_image_info(image_dir: Union[str, Path]) -> List[Dict[str, Any]]:
    image_dir = Path(image_dir).expanduser().resolve()
    if not image_dir.exists():
        raise FileNotFoundError(f"Directory does not exist: {image_dir}")
    if not image_dir.is_dir():
        raise NotADirectoryError(f"Not a directory: {image_dir}")

    rows: List[Dict[str, Any]] = []
    for img_path in sorted(image_dir.iterdir()):
        if not img_path.is_file():
            continue
        if img_path.suffix.lower() not in IMAGE_SUFFIXES:
            continue
        if ".warm." in img_path.name:
            continue
        m = FILENAME_PATTERN.match(img_path.name)
        if not m:
            continue

        info = m.groupdict()
        model = info["model"]
        prompt_no = info["prompt_no"]
        language = info["language"].lower()
        width = int(info["width"])
        height = int(info["height"])
        step = int(info["step"])
        seed_no = info["seed_no"]

        txt_path = img_path.with_suffix(".txt")
        if not txt_path.exists():
            raise FileNotFoundError(f"Missing corresponding txt file: {txt_path}")
        txt_info = _read_sidecar_txt(txt_path)

        prompt_text = _pick_first_value(txt_info, ["prompt_text", "prompt", "text"], default=None)
        if not prompt_text:
            raise ValueError(f"prompt_text/prompt was not found in {txt_path}")

        elapsed_seconds = _to_float(_get_nested(txt_info, "monitor.elapsed_seconds"))
        if elapsed_seconds is None:
            elapsed_seconds = _to_float(
                _pick_first_value(
                    txt_info,
                    ["elapsed_seconds", "elapsed", "time_seconds", "seconds", "duration", "ElapsedSeconds"],
                    default=None,
                )
            )

        iter_per_second = None
        if elapsed_seconds is not None and elapsed_seconds > 0:
            iter_per_second = step / elapsed_seconds

        vram_usage, gpu_usage, memory_usage, cpu_usage = _extract_resource_fields(txt_info)

        row: Dict[str, Any] = {
            "model": model,
            "prompt_no": prompt_no,
            "language": language,
            "resolution": f"{width}*{height}",
            "step": step,
            "seed_no": seed_no,
            "image_size_bytes": img_path.stat().st_size,
            "elapsed_seconds": elapsed_seconds,
            "iter_per_second": iter_per_second,
            "vram_usage": vram_usage,
            "gpu_usage": gpu_usage,
            "memory_usage": memory_usage,
            "cpu_usage": cpu_usage,
            "filename": img_path.name,
        }
        for metric in METRIC_COLUMNS:
            row[metric] = None
        rows.append(row)
    return rows


def _iter_jsonl(path: Union[str, Path]) -> Iterable[Dict[str, Any]]:
    jsonl_path = Path(path).expanduser().resolve()
    with jsonl_path.open("r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, start=1):
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"Invalid JSONL line {line_no} in {jsonl_path}: {exc}") from exc
            if isinstance(obj, dict):
                yield obj


def load_score_jsonl(score_jsonl_paths: Sequence[Union[str, Path]]) -> Tuple[Dict[str, Dict[str, Any]], List[Dict[str, Any]]]:
    score_map: Dict[str, Dict[str, Any]] = {}
    error_rows: List[Dict[str, Any]] = []
    for path in score_jsonl_paths:
        for rec in _iter_jsonl(path):
            filename = rec.get("filename")
            if not filename:
                continue
            score_map.setdefault(str(filename), {})
            scores = rec.get("scores", {})
            if isinstance(scores, dict):
                for metric, value in scores.items():
                    if metric in METRIC_COLUMNS:
                        score_map[str(filename)][metric] = value
            errors = rec.get("errors", {})
            if isinstance(errors, dict):
                for metric, message in errors.items():
                    if message:
                        error_rows.append({
                            "source_jsonl": str(path),
                            "filename": filename,
                            "metric": metric,
                            "error": str(message),
                        })
    return score_map, error_rows


def load_metric_resource_rows(score_jsonl_paths: Sequence[Union[str, Path]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for path in score_jsonl_paths:
        for rec in _iter_jsonl(path):
            filename = rec.get("filename")
            scores = rec.get("scores", {})
            errors = rec.get("errors", {})
            measurements = rec.get("measurements", {})
            if not filename or not isinstance(measurements, dict):
                continue
            for metric, measurement in measurements.items():
                if metric not in METRIC_COLUMNS or not isinstance(measurement, dict):
                    continue
                rows.append({
                    "filename": filename,
                    "metric": metric,
                    "score": scores.get(metric) if isinstance(scores, dict) else None,
                    "elapsed_seconds": measurement.get("elapsed_seconds"),
                    "started_at_utc": measurement.get("started_at_utc"),
                    "ended_at_utc": measurement.get("ended_at_utc"),
                    "sample_interval_seconds": measurement.get("sample_interval_seconds"),
                    "sample_count": measurement.get("sample_count"),
                    "process_cpu_avg": _get_nested(measurement, "process.cpu_percent.avg"),
                    "process_cpu_max": _get_nested(measurement, "process.cpu_percent.max"),
                    "process_memory_avg": _get_nested(measurement, "process.memory_rss_mb.avg"),
                    "process_memory_max": _get_nested(measurement, "process.memory_rss_mb.max"),
                    "system_cpu_avg": _get_nested(measurement, "system_cpu_percent.avg"),
                    "system_cpu_max": _get_nested(measurement, "system_cpu_percent.max"),
                    "system_memory_percent_avg": _get_nested(measurement, "system_memory_percent.avg"),
                    "system_memory_percent_max": _get_nested(measurement, "system_memory_percent.max"),
                    "system_memory_used_avg": _get_nested(measurement, "system_memory_used_mb.avg"),
                    "system_memory_used_max": _get_nested(measurement, "system_memory_used_mb.max"),
                    "gpu_index": _get_nested(measurement, "gpu.index"),
                    "gpu_name": _get_nested(measurement, "gpu.name"),
                    "gpu_utilization_avg": _get_nested(measurement, "gpu.utilization_percent.avg"),
                    "gpu_utilization_max": _get_nested(measurement, "gpu.utilization_percent.max"),
                    "gpu_memory_avg": _get_nested(measurement, "gpu.memory_used_mb.avg"),
                    "gpu_memory_max": _get_nested(measurement, "gpu.memory_used_mb.max"),
                    "monitor_error": measurement.get("monitor_error"),
                    "metric_error": errors.get(metric) if isinstance(errors, dict) else None,
                })
    return rows


def merge_scores(rows: List[Dict[str, Any]], score_jsonl_paths: Sequence[Union[str, Path]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    if not score_jsonl_paths:
        return rows, []
    score_map, error_rows = load_score_jsonl(score_jsonl_paths)
    for row in rows:
        filename = row.get("filename")
        if filename in score_map:
            row.update(score_map[filename])
    return rows, error_rows


def _set_column_widths(ws: Worksheet) -> None:
    default_widths = {
        "ModelShortName": 14,
        "PromptID": 12,
        "Language": 12,
        "Resolution": 14,
        "Steps": 8,
        "SeedNo": 14,
        "ImageSizeBytes": 18,
        "ElapsedSeconds": 16,
        "IterPerSecond": 18,
        "VRAMUsage": 16,
        "GPUUsage": 14,
        "MemoryUsage": 14,
        "CPUUsage": 12,
        "FileName": 54,
    }
    for col_idx, (_, header_name) in enumerate(OUTPUT_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = default_widths.get(header_name, 18)


def save_results_to_excel(rows: List[Dict[str, Any]], output_path: Union[str, Path], error_rows: Optional[List[Dict[str, Any]]] = None) -> Path:
    output_path = Path(output_path).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "image_scores"
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(OUTPUT_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{max(1, len(rows) + 1)}"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    top_alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, (_, header_name) in enumerate(OUTPUT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    metric_key_to_col: Dict[str, int] = {}
    for col_idx, (key, _) in enumerate(OUTPUT_HEADERS, start=1):
        if key in METRIC_COLUMNS:
            metric_key_to_col[key] = col_idx

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(OUTPUT_HEADERS, start=1):
            value = row_data.get(key)
            if key == "seed_no" and value is not None:
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                cell.number_format = "@"
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = top_alignment

    for row_idx in range(2, len(rows) + 2):
        for col_idx in metric_key_to_col.values():
            ws.cell(row=row_idx, column=col_idx).number_format = "0.0000"
        for header_name in ["ElapsedSeconds", "IterPerSecond"]:
            col_idx = [h for _, h in OUTPUT_HEADERS].index(header_name) + 1
            ws.cell(row=row_idx, column=col_idx).number_format = "0.000"

    ws.row_dimensions[1].height = 22
    _set_column_widths(ws)

    note_ws = wb.create_sheet("Notes")
    notes = [
        ("FieldNotes", ""),
        ("Metric columns", "OpenCLIPScore, CNCLIPScore, AestheticScore, MUSIQScore, ImageRewardScore, HPSv2.1Score and VQAScore are merged from JSONL files generated in different conda environments."),
        ("IterPerSecond", "Number of sampling steps executed per second. It is calculated as steps / elapsed_seconds."),
        ("VRAMUsage", "Peak VRAM usage is used by default. If the peak VRAM field is not available in the txt file, the generic VRAM field is used instead."),
        ("GPUUsage / MemoryUsage / CPUUsage", "Average utilization is used by default. If the average value field is not available in the txt file, the generic field is used instead."),
        ("SeedNo", "Parsed from the image filename and written to Excel as text to avoid automatic scientific notation."),
        ("FileName", "Only the base filename is retained, without the path."),
    ]
    for row_idx, (key, value) in enumerate(notes, start=1):
        note_ws.cell(row=row_idx, column=1, value=key)
        note_ws.cell(row=row_idx, column=2, value=value)
    note_ws["A1"].font = Font(bold=True)
    note_ws.column_dimensions["A"].width = 34
    note_ws.column_dimensions["B"].width = 100
    for row in note_ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    if error_rows:
        err_ws = wb.create_sheet("MetricErrors")
        err_headers = ["SourceJSONL", "FileName", "Metric", "Error"]
        for col_idx, header in enumerate(err_headers, start=1):
            cell = err_ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        for row_idx, err in enumerate(error_rows, start=2):
            err_ws.cell(row=row_idx, column=1, value=err.get("source_jsonl"))
            err_ws.cell(row=row_idx, column=2, value=err.get("filename"))
            err_ws.cell(row=row_idx, column=3, value=err.get("metric"))
            err_ws.cell(row=row_idx, column=4, value=err.get("error"))
        err_ws.freeze_panes = "A2"
        err_ws.auto_filter.ref = f"A1:D{max(1, len(error_rows) + 1)}"
        err_ws.column_dimensions["A"].width = 42
        err_ws.column_dimensions["B"].width = 54
        err_ws.column_dimensions["C"].width = 20
        err_ws.column_dimensions["D"].width = 120
        for row in err_ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_path)
    return output_path


def save_metric_resources_to_excel(rows: List[Dict[str, Any]], output_path: Union[str, Path]) -> Path:
    output_path = Path(output_path).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "metric_resources"
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(RESOURCE_OUTPUT_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{max(1, len(rows) + 1)}"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, (_, header) in enumerate(RESOURCE_OUTPUT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(RESOURCE_OUTPUT_HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if key == "score":
                cell.number_format = "0.0000"
            elif key in {"elapsed_seconds", "sample_interval_seconds"}:
                cell.number_format = "0.0000"
            elif key.endswith("_avg") or key.endswith("_max"):
                cell.number_format = "0.000"

    for col_idx, (_, header) in enumerate(RESOURCE_OUTPUT_HEADERS, start=1):
        if header == "FileName":
            width = 54
        elif header in {"StartedAtUTC", "EndedAtUTC"}:
            width = 30
        elif header in {"MonitorError", "MetricError"}:
            width = 50
        else:
            width = max(14, min(28, len(header) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    note_ws = wb.create_sheet("Notes")
    notes = [
        ("FieldNotes", ""),
        ("Row granularity", "Each row represents one image and one objective metric computation."),
        ("ElapsedSeconds", "Wall-clock time around the metric call. The first call may include lazy model initialization."),
        ("Process fields", "CPU and RSS memory include the scoring Python process and its child processes."),
        ("GPU fields", "GPU utilization and memory are device-wide values sampled with nvidia-smi, not process-attributed values."),
        ("Sampling", "Sampling follows resource_monitor.py. Short computations may contain only one sample."),
    ]
    for row_idx, (key, value) in enumerate(notes, start=1):
        note_ws.cell(row=row_idx, column=1, value=key)
        note_ws.cell(row=row_idx, column=2, value=value)
    note_ws["A1"].font = Font(bold=True)
    note_ws.column_dimensions["A"].width = 24
    note_ws.column_dimensions["B"].width = 110
    for row in note_ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_path)
    return output_path


def process_image_dir_to_excel(
    image_dir: Union[str, Path],
    score_jsonl_paths: Optional[Sequence[Union[str, Path]]] = None,
    output_filename: str = "image_info_scores_all_metrics.xlsx",
    resource_output_filename: Optional[str] = None,
) -> Path:
    image_dir = Path(image_dir).expanduser().resolve()
    rows = extract_image_info(image_dir=image_dir)
    rows, error_rows = merge_scores(rows, score_jsonl_paths or [])
    output_path = image_dir / output_filename
    saved_path = save_results_to_excel(rows, output_path, error_rows=error_rows)
    if score_jsonl_paths:
        resource_rows = load_metric_resource_rows(score_jsonl_paths)
        resource_name = resource_output_filename or f"{Path(output_filename).stem}_metric_resources.xlsx"
        resource_output_path = image_dir / resource_name
        if resource_output_path.resolve() == output_path.resolve():
            raise ValueError("Resource Excel output must differ from the score Excel output.")
        save_metric_resources_to_excel(resource_rows, resource_output_path)
    return saved_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Extract image metadata and merge objective metrics into Excel.")
    parser.add_argument("image_dir", help="Image directory.")
    parser.add_argument(
        "--score-jsonl",
        action="append",
        default=[],
        help="Metric JSONL file produced by objective_score.py. Can be used multiple times.",
    )
    parser.add_argument(
        "--score-jsonl-glob",
        default=None,
        help="Optional glob pattern for metric JSONL files, e.g. .objective_scores_tmp/*.jsonl.",
    )
    parser.add_argument(
        "--output-filename",
        default="image_info_scores_all_metrics.xlsx",
        help="Excel filename written under image_dir.",
    )
    parser.add_argument(
        "--resource-output-filename",
        default=None,
        help="Resource Excel filename under image_dir. Default: <score filename stem>_metric_resources.xlsx.",
    )
    args = parser.parse_args()

    score_paths: List[Path] = [Path(p).expanduser().resolve() for p in args.score_jsonl]
    if args.score_jsonl_glob:
        score_paths.extend(sorted(Path().glob(args.score_jsonl_glob)))

    output_path = process_image_dir_to_excel(
        image_dir=args.image_dir,
        score_jsonl_paths=score_paths,
        output_filename=args.output_filename,
        resource_output_filename=args.resource_output_filename,
    )
    print(f"Excel: {output_path}")
    resource_name = args.resource_output_filename or f"{Path(args.output_filename).stem}_metric_resources.xlsx"
    print(f"Metric resources Excel: {Path(args.image_dir).expanduser().resolve() / resource_name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
