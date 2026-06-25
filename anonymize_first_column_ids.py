# anonymize_first_column_ids.py
# 用法示例：
# python anonymize_first_column_ids.py input.xlsx output_anonymized.xlsx
# 如果第一行不是表头：
# python anonymize_first_column_ids.py input.xlsx output_anonymized.xlsx --header-row 0
# 如果需要导出原ID与匿名ID映射表：
# python anonymize_first_column_ids.py input.xlsx output_anonymized.xlsx --mapping-out id_mapping.csv

import argparse
import csv
from pathlib import Path
from openpyxl import load_workbook


def normalize_id(value):
    """
    将单元格中的 ID 统一转换成字符串，用于建立映射。
    空值返回 None，不进行匿名化。
    """
    if value is None:
        return None

    text = str(value).strip()
    if text == "":
        return None

    return text


def anonymize_excel(input_path, output_path, header_row=1, prefix="R", width=4, mapping_out=None):
    input_path = Path(input_path)
    output_path = Path(output_path)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    # keep_vba=True 可尽量保留 xlsm 宏文件结构；普通 xlsx 也可用
    wb = load_workbook(input_path, keep_vba=input_path.suffix.lower() == ".xlsm")

    id_mapping = {}
    next_index = 1

    # 第一步：扫描所有 sheet 的第一列，建立全局映射
    for ws in wb.worksheets:
        start_row = header_row + 1 if header_row and header_row > 0 else 1

        for row in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            original_id = normalize_id(cell.value)

            if original_id is None:
                continue

            if original_id not in id_mapping:
                anonymized_id = f"{prefix}{next_index:0{width}d}"
                id_mapping[original_id] = anonymized_id
                next_index += 1

    # 第二步：替换所有 sheet 第一列中的 ID
    for ws in wb.worksheets:
        start_row = header_row + 1 if header_row and header_row > 0 else 1

        for row in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            original_id = normalize_id(cell.value)

            if original_id is None:
                continue

            cell.value = id_mapping[original_id]

    # 保存匿名化后的 Excel
    wb.save(output_path)

    # 可选：导出映射表。注意：这个文件不要提交给期刊或公开。
    if mapping_out:
        mapping_out = Path(mapping_out)
        with mapping_out.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["OriginalID", "AnonymizedID"])
            for original_id, anonymized_id in id_mapping.items():
                writer.writerow([original_id, anonymized_id])

    print(f"Done.")
    print(f"Input file: {input_path}")
    print(f"Output file: {output_path}")
    print(f"Unique respondent IDs anonymized: {len(id_mapping)}")

    if mapping_out:
        print(f"Mapping file saved to: {mapping_out}")
        print("Do NOT upload or publicly share the mapping file.")


def main():
    parser = argparse.ArgumentParser(
        description="Anonymize respondent IDs in the first column of every sheet in an Excel workbook."
    )

    parser.add_argument("input", help="Input Excel file, e.g., input.xlsx")
    parser.add_argument("output", help="Output anonymized Excel file, e.g., output_anonymized.xlsx")

    parser.add_argument(
        "--header-row",
        type=int,
        default=1,
        help="Header row number. Default is 1. Use 0 if there is no header row."
    )

    parser.add_argument(
        "--prefix",
        default="R",
        help="Prefix for anonymized IDs. Default is R, producing IDs such as R0001."
    )

    parser.add_argument(
        "--width",
        type=int,
        default=4,
        help="Number width for anonymized IDs. Default is 4, producing IDs such as R0001."
    )

    parser.add_argument(
        "--mapping-out",
        default=None,
        help="Optional CSV file to save the mapping between original IDs and anonymized IDs."
    )

    args = parser.parse_args()

    anonymize_excel(
        input_path=args.input,
        output_path=args.output,
        header_row=args.header_row,
        prefix=args.prefix,
        width=args.width,
        mapping_out=args.mapping_out
    )


if __name__ == "__main__":
    main()